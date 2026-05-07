import streamlit as st
import requests
import pandas as pd
from datetime import datetime, timedelta
import plotly.express as px
from io import BytesIO
import re
from typing import List, Dict, Any

# Configuração da página
st.set_page_config(
    page_title="Monitor PNCP - Evolutio",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Constantes
API_BASE_URL = "https://pncp.gov.br/api/pncp/v1/contratacoes"
ESTADOS = ["RO", "AC", "MT", "AM"]
MODALIDADE = 6  # Pregão Eletrônico
PALAVRAS_CHAVE = [
    "limpeza", "conservação", "serviços gerais", "portaria", "recepção", 
    "jardinagem", "apoio administrativo", "auxiliar administrativo", 
    "secretaria", "copeira", "zeladoria", "vigilância desarmada", "brigadista"
]
PALAVRAS_EXCLUSAO = ["armada", "compra de material", "material de construção"]

@st.cache_data(ttl=1800)  # Cache de 30 minutos
def buscar_contratacoes(**filtros):
    """Busca contratacoes na API do PNCP com filtros aplicados"""
    params = {
        "modalidade": MODALIDADE,
        "situacao": "ativa",
        "page": 1,
        "size": 100
    }
    
    # Adiciona filtros passados
    for key, value in filtros.items():
        if value:
            params[key] = value
    
    try:
        response = requests.get(API_BASE_URL, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        return data.get("content", [])
    except Exception as e:
        st.error(f"Erro ao consultar API: {str(e)}")
        return []

def filtrar_por_estado(contratacoes: List[Dict]) -> List[Dict]:
    """Filtra apenas pelos estados RO, AC, MT, AM"""
    return [c for c in contratacoes if c.get("orgao_uf", "").upper() in ESTADOS]

def filtrar_por_data(contratacoes: List[Dict], horas: int = 24) -> List[Dict]:
    """Filtra por data de publicação recente"""
    cutoff = datetime.now() - timedelta(hours=horas)
    return [c for c in contratacoes if 
            datetime.fromisoformat(c.get("data_publicacao", "").replace('Z', '+00:00')) > cutoff]

def contem_palavra_chave(objeto: str) -> bool:
    """Verifica se o objeto contém alguma palavra-chave relevante"""
    objeto_lower = objeto.lower()
    return any(palavra in objeto_lower for palavra in PALAVRAS_CHAVE)

def excluir_palavras(objeto: str) -> bool:
    """Exclui resultados com palavras de exclusão"""
    objeto_lower = objeto.lower()
    return any(palavra in objeto_lower for palavra in PALAVRAS_EXCLUSAO)

def processar_dados(contratacoes: List[Dict]) -> pd.DataFrame:
    """Processa os dados para exibição na tabela"""
    dados_processados = []
    
    for c in contratacoes:
        objeto = c.get("objeto", "")
        
        # Aplica filtros de palavras-chave e exclusão
        if not contem_palavra_chave(objeto) or excluir_palavras(objeto):
            continue
            
        try:
            valor_estimado = float(c.get("valor_estimado", 0))
        except:
            valor_estimado = 0
            
        dados_processados.append({
            "Órgão": c.get("orgao_nome", ""),
            "Localidade": f"{c.get('orgao_municipio', '')} - {c.get('orgao_uf', '')}",
            "Objeto": objeto[:200] + "..." if len(objeto) > 200 else objeto,
            "Valor Estimado": f"R$ {valor_estimado:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            "Link Edital": c.get("url_ativa", ""),
            "Data Publicação": c.get("data_publicacao", ""),
            "Valor Numérico": valor_estimado,
            "UF": c.get("orgao_uf", ""),
            "CNPJ": c.get("orgao_cnpj", "")
        })
    
    return pd.DataFrame(dados_processados)

# Sidebar
st.sidebar.title("🔧 Filtros")
filtro_recencia = st.sidebar.slider("Últimas horas", 1, 168, 24)  # Até 7 dias
limite_valor_alerta = st.sidebar.number_input("Alerta acima de (R$)", value=500000.0, step=50000.0)
cnpj_busca = st.sidebar.text_input("Buscar por CNPJ do Órgão")

if st.sidebar.button("🔄 Atualizar Dados"):
    st.cache_data.clear()
    st.rerun()

# Título principal
st.title("📋 Monitor PNCP - Evolutio")
st.markdown("---")

# Busca principal
with st.spinner("Buscando pregões eletrônicos nos estados RO, AC, MT, AM..."):
    contratacoes = buscar_contratacoes()
    contratacoes_filtradas = filtrar_por_estado(contratacoes)
    
    if filtro_recencia:
        contratacoes_filtradas = filtrar_por_data(contratacoes_filtradas, filtro_recencia)
    
    df = processar_dados(contratacoes_filtradas)

# Filtro por CNPJ
if cnpj_busca:
    df = df[df["CNPJ"].str.contains(cnpj_busca, na=False)]

# Métricas principais
col1, col2, col3, col4 = st.columns(4)
total_encontrados = len(df)
col1.metric("📊 Total Encontrados", total_encontrados)
col2.metric("🔥 Recentes (24h)", len(df[df["Data Publicação"].str.contains("T", na=False)]))
col3.metric("⚠️ Alto Valor", len(df[df["Valor Numérico"] > limite_valor_alerta]))
col4.metric("📈 Estados Ativos", df["UF"].nunique())

# Dashboard de Estados
if not df.empty:
    st.subheader("📈 Dashboard por Estado")
    fig = px.bar(
        df["UF"].value_counts().reset_index(),
        x="UF", y="count", 
        title="Distribuição de Editais por Estado",
        color="UF",
        color_discrete_sequence=px.colors.qualitative.Set3
    )
    st.plotly_chart(fig, use_container_width=True)

# Tabela principal
st.subheader("📋 Editais Encontrados")

if df.empty:
    st.warning("Nenhum edital encontrado com os filtros aplicados.")
else:
    # Formatação condicional
    def highlight_alto_valor(row):
        if row["Valor Numérico"] > limite_valor_alerta:
            return ["background-color: #ffebee"] * len(row)
        return [""] * len(row)
    
    styled_df = df.style.apply(highlight_alto_valor, axis=1)
    
    # Configuração da tabela
    st.dataframe(
        styled_df,
        column_config={
            "Link Edital": st.column_config.LinkColumn("🔗 Acessar Edital"),
            "Valor Estimado": st.column_config.Column("💰 Valor", width="150px"),
            "Valor Numérico": st.column_config.Column("Valor (invisível)", disabled=True),
            "CNPJ": st.column_config.Column("CNPJ", disabled=True),
            "Data Publicação": st.column_config.Column("Data", disabled=True),
            "UF": st.column_config.Column("UF", disabled=True)
        },
        hide_index=True,
        use_container_width=True
    )

# Exportação Excel
if not df.empty:
    col1, col2 = st.columns([3, 1])
    with col2:
        def export_excel():
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_full = df.drop(columns=["Valor Numérico", "UF", "CNPJ"])
                df_full.to_excel(writer, sheet_name='Editais PNCP', index=False)
                
                # Formatação
                worksheet = writer.sheets['Editais PNCP']
                from openpyxl.styles import PatternFill, Font
                for row in worksheet.iter_rows(min_row=2, max_col=5):
                    valor = float(row[3].value.replace('R$ ', '').replace('.', '').replace(',', '.'))
                    if valor > limite_valor_alerta:
                        for cell in row:
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            output.seek(0)
            st.download_button(
                label="📥 Exportar Excel",
                data=output.getvalue(),
                file_name=f"pncp_evolutio_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        export_excel()

# Footer
st.markdown("---")
st.markdown("""
**Evolutio Monitor PNCP**  
*Desenvolvido para monitoramento automático de licitações de serviços*  
🔄 Atualização automática a cada 30 minutos | 📱 Otimizado para mobile
""")

# Auto-refresh opcional
if st.button("🔄 Refresh Automático (30s)"):
    st.rerun()